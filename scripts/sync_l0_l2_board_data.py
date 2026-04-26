from __future__ import annotations

import json
import math
import http.client
import re
import ssl
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = Path("/Users/moonlanner/Desktop/stocks analysis/L0-L2stocks.xlsx")
OUTPUT_PATH = ROOT / "src" / "industry-board" / "generatedIndustryData.js"
FUTU_HOST = "127.0.0.1"
FUTU_PORT = 11111
HTTP_TIMEOUT_SECONDS = 10

LAYER_META = {
    "L0": {"id": "l0", "name": "能源层", "nameEn": "Energy"},
    "L1": {"id": "l1", "name": "芯片层", "nameEn": "Semiconductors"},
    "L2": {"id": "l2", "name": "基础设施层", "nameEn": "Infrastructure"},
}

CN_NUM_MAP = {
    "一": 1,
    "二": 2,
    "三": 3,
    "四": 4,
    "五": 5,
    "六": 6,
    "七": 7,
    "八": 8,
    "九": 9,
    "十": 10,
}


def parse_number(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        number = float(value)
        return number if math.isfinite(number) else None

    text = str(value).strip().replace(",", "")
    if not text:
        return None

    try:
        number = float(text)
        return number if math.isfinite(number) else None
    except ValueError:
        return None


def is_finite_number(value) -> bool:
    return isinstance(value, (int, float)) and math.isfinite(value)


def slugify(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")
    return slug or "sector"


def normalize_rank_label(raw_label: str) -> str:
    text = raw_label.strip()
    if text.isdigit():
        return text

    if "_" in text:
        return text.replace("_", "-")

    if text in CN_NUM_MAP:
        return str(CN_NUM_MAP[text])

    return text


def prettify_sector_name(raw_name: str) -> str:
    text = raw_name.replace("_", " / ").strip()
    text = re.sub(r"([\u4e00-\u9fff])([A-Za-z])", r"\1 \2", text)
    text = re.sub(r"([A-Za-z])([\u4e00-\u9fff])", r"\1 \2", text)
    text = re.sub(r"\s+", " ", text).strip()

    def _title_token(match):
        token = match.group(0)
        return token if token.isupper() else token.title()

    return re.sub(r"[A-Za-z]+(?: [A-Za-z]+)*", _title_token, text)


def parse_sheet_meta(sheet_name: str):
    match = re.match(r"^(L\d)细分([0-9_]+|[一二三四五六七八九十]+)\s*(.*)$", sheet_name.strip())
    if not match:
        raise ValueError(f"Unrecognized sheet name format: {sheet_name}")

    layer_code, raw_rank, raw_name = match.groups()
    sector_name = prettify_sector_name(raw_name or sheet_name)
    rank_label = normalize_rank_label(raw_rank)
    sector_code = f"{layer_code}T{rank_label}"
    sector_id = f"{layer_code.lower()}-{slugify(rank_label)}-{slugify(sector_name)}"
    return layer_code, sector_code, sector_name, sector_id


def normalize_ticker(raw_ticker: str) -> str:
    market, _, symbol = raw_ticker.partition(".")
    market = market.upper()
    if not symbol:
        return raw_ticker

    if market == "US":
        return symbol
    if market in {"SH", "SZ", "HK", "KS", "BJ"}:
        return f"{symbol}.{market}"
    return raw_ticker


def build_quote(price, open_price, high, low, prev_close, name=None, source=None, as_of=None, volume=None, turnover=None):
    values = {
        "price": price,
        "open": open_price,
        "high": high,
        "low": low,
        "prevClose": prev_close,
        "volume": volume,
        "turnover": turnover,
    }
    quote = {
        key: float(value)
        for key, value in values.items()
        if is_finite_number(value)
    }
    if name:
        quote["name"] = str(name).strip()
    if source:
        quote["source"] = source
    if as_of:
        quote["asOf"] = str(as_of)
    if "price" in quote and "prevClose" in quote and quote["prevClose"] != 0:
        quote["dailyChange"] = ((quote["price"] - quote["prevClose"]) / quote["prevClose"]) * 100
    return quote


def build_snapshot_trend(price, open_price, high, low, prev_close):
    if price is None or prev_close in (None, 0):
        return []

    points = [
        ("Prev", prev_close),
        ("Open", open_price),
        ("Low", low),
        ("High", high),
        ("Last", price),
    ]

    return [
        {"label": label, "value": (value / prev_close) * 100}
        for label, value in points
        if value is not None
    ]


def average_company_trends(companies):
    trends = [company["trend"] for company in companies if company.get("trend")]
    if not trends:
        return []

    labels = max(trends, key=len)
    averaged = []
    for index, point in enumerate(labels):
        values = [
            trend[index]["value"]
            for trend in trends
            if len(trend) > index and trend[index].get("value") is not None
        ]
        if values:
            averaged.append({"label": point["label"], "value": sum(values) / len(values)})

    return averaged


def futu_row_to_quote(row):
    return build_quote(
        price=parse_number(row.get("last_price")),
        open_price=parse_number(row.get("open_price")),
        high=parse_number(row.get("high_price")),
        low=parse_number(row.get("low_price")),
        prev_close=parse_number(row.get("prev_close_price")),
        name=row.get("name"),
        source="futu",
        as_of=row.get("update_time"),
        volume=parse_number(row.get("volume")),
        turnover=parse_number(row.get("turnover")),
    )


def fetch_futu_quotes(tickers):
    try:
        from futu import OpenQuoteContext, RET_OK
    except Exception as exc:
        print(f"Futu SDK unavailable, using workbook/fallback only: {exc}", file=sys.stderr)
        return {}, {ticker: "futu sdk unavailable" for ticker in tickers}

    quotes = {}
    failures = {}
    ctx = OpenQuoteContext(host=FUTU_HOST, port=FUTU_PORT)

    def query_batch(batch):
        if not batch:
            return
        ret, data = ctx.get_market_snapshot(batch)
        if ret == RET_OK:
            for row in data.to_dict("records"):
                code = row.get("code")
                if code:
                    quotes[code] = futu_row_to_quote(row)
            return

        if len(batch) == 1:
            failures[batch[0]] = str(data)
            return

        mid = len(batch) // 2
        query_batch(batch[:mid])
        query_batch(batch[mid:])

    try:
        for start in range(0, len(tickers), 50):
            query_batch(tickers[start:start + 50])
    except Exception as exc:
        for ticker in tickers:
            if ticker not in quotes:
                failures.setdefault(ticker, str(exc))
    finally:
        ctx.close()

    return quotes, failures


def fetch_text(url, headers=None):
    req = urllib.request.Request(url, headers=headers or {})
    context = None
    try:
        import certifi

        context = ssl.create_default_context(cafile=certifi.where())
    except Exception:
        context = ssl.create_default_context()

    last_exc = None
    for _ in range(2):
        try:
            with urllib.request.urlopen(req, timeout=HTTP_TIMEOUT_SECONDS, context=context) as response:
                return response.read()
        except http.client.IncompleteRead as exc:
            last_exc = exc
            time.sleep(0.2)
    raise last_exc


def fetch_yahoo_quote(ticker: str):
    market, _, symbol = ticker.partition(".")
    if market != "US" or not symbol:
        return None

    url = f"https://query2.finance.yahoo.com/v8/finance/chart/{urllib.parse.quote(symbol)}?range=1d&interval=1d"
    try:
        raw = fetch_text(url, headers={"User-Agent": "Mozilla/5.0"})
        data = json.loads(raw.decode("utf-8"))
        result = (data.get("chart", {}).get("result") or [None])[0]
        meta = result.get("meta") if result else None
        if not meta:
            return None
        as_of = None
        if meta.get("regularMarketTime"):
            as_of = datetime.fromtimestamp(meta["regularMarketTime"]).astimezone().isoformat(timespec="seconds")
        return build_quote(
            price=parse_number(meta.get("regularMarketPrice")),
            open_price=parse_number(meta.get("regularMarketOpen")),
            high=parse_number(meta.get("regularMarketDayHigh")),
            low=parse_number(meta.get("regularMarketDayLow")),
            prev_close=parse_number(meta.get("chartPreviousClose") or meta.get("previousClose")),
            name=meta.get("longName") or meta.get("shortName"),
            source="yahoo",
            as_of=as_of,
            volume=parse_number(meta.get("regularMarketVolume")),
        )
    except (urllib.error.URLError, TimeoutError, http.client.IncompleteRead, json.JSONDecodeError, KeyError, TypeError) as exc:
        print(f"Yahoo fallback failed for {ticker}: {exc}", file=sys.stderr)
        return None


def yahoo_chart_symbol(ticker: str):
    market, _, symbol = ticker.partition(".")
    if not symbol:
        return ticker
    if market == "US":
        return symbol
    if market == "HK":
        return f"{symbol.zfill(4)}.HK"
    if market == "SH":
        return f"{symbol}.SS"
    if market == "SZ":
        return f"{symbol}.SZ"
    return None


def fetch_yahoo_monthly_trend(ticker: str):
    symbol = yahoo_chart_symbol(ticker)
    if not symbol:
        return []

    url = (
        f"https://query2.finance.yahoo.com/v8/finance/chart/{urllib.parse.quote(symbol)}"
        "?range=7mo&interval=1mo&includePrePost=false"
    )
    try:
        raw = fetch_text(url, headers={"User-Agent": "Mozilla/5.0"})
        data = json.loads(raw.decode("utf-8"))
        result = (data.get("chart", {}).get("result") or [None])[0]
        if not result:
            return []

        timestamps = result.get("timestamp") or []
        quote = ((result.get("indicators") or {}).get("quote") or [{}])[0]
        opens = quote.get("open") or []
        highs = quote.get("high") or []
        lows = quote.get("low") or []
        closes = quote.get("close") or []
        points = []

        for index, timestamp in enumerate(timestamps):
            open_price = parse_number(opens[index] if index < len(opens) else None)
            high = parse_number(highs[index] if index < len(highs) else None)
            low = parse_number(lows[index] if index < len(lows) else None)
            close = parse_number(closes[index] if index < len(closes) else None)
            if not all(is_finite_number(value) for value in (open_price, high, low, close)) or open_price == 0:
                continue

            month = datetime.fromtimestamp(timestamp).astimezone().strftime("%Y-%m")
            points.append({
                "month": month,
                "label": month[5:],
                "open": 0.0,
                "high": ((high - open_price) / open_price) * 100,
                "low": ((low - open_price) / open_price) * 100,
                "close": ((close - open_price) / open_price) * 100,
            })

        deduped = {point["month"]: point for point in points}
        return [deduped[month] for month in sorted(deduped)][-6:]
    except (urllib.error.URLError, TimeoutError, http.client.IncompleteRead, json.JSONDecodeError, KeyError, TypeError) as exc:
        print(f"Yahoo monthly trend failed for {ticker}: {exc}", file=sys.stderr)
        return []


def sina_code(ticker: str):
    market, _, symbol = ticker.partition(".")
    if not symbol:
        return None
    if market == "US":
        return f"gb_{symbol.lower()}"
    if market == "HK":
        return f"hk{symbol.zfill(5)}"
    if market in {"SH", "SZ"}:
        if symbol.startswith(("4", "8", "9")):
            return f"bj{symbol}"
        return f"{market.lower()}{symbol}"
    return None


def fetch_sina_quote(ticker: str):
    code = sina_code(ticker)
    if not code:
        return None

    url = f"https://hq.sinajs.cn/list={urllib.parse.quote(code)}"
    try:
        raw = fetch_text(url, headers={
            "Referer": "https://finance.sina.com.cn",
            "User-Agent": "Mozilla/5.0",
        })
        text = raw.decode("gb18030", errors="ignore")
        match = re.search(r'"([^"]*)"', text)
        if not match or not match.group(1):
            return None
        parts = match.group(1).split(",")
        if code.startswith("gb_"):
            price = parse_number(parts[1] if len(parts) > 1 else None)
            change = parse_number(parts[4] if len(parts) > 4 else None)
            prev_close = parse_number(parts[26] if len(parts) > 26 else None)
            if prev_close is None and price is not None and change is not None:
                prev_close = price - change
            return build_quote(
                price=price,
                open_price=parse_number(parts[5] if len(parts) > 5 else None),
                high=parse_number(parts[6] if len(parts) > 6 else None),
                low=parse_number(parts[7] if len(parts) > 7 else None),
                prev_close=prev_close,
                name=parts[0] if parts else None,
                source="sina",
                as_of=parts[3] if len(parts) > 3 else None,
                volume=parse_number(parts[10] if len(parts) > 10 else None),
            )
        if code.startswith("hk"):
            return build_quote(
                price=parse_number(parts[6] if len(parts) > 6 else None),
                open_price=parse_number(parts[2] if len(parts) > 2 else None),
                high=parse_number(parts[4] if len(parts) > 4 else None),
                low=parse_number(parts[5] if len(parts) > 5 else None),
                prev_close=parse_number(parts[3] if len(parts) > 3 else None),
                name=parts[1] if len(parts) > 1 else None,
                source="sina",
                as_of=f"{parts[17]} {parts[18]}" if len(parts) > 18 else None,
                volume=parse_number(parts[12] if len(parts) > 12 else None),
                turnover=parse_number(parts[11] if len(parts) > 11 else None),
            )
        return build_quote(
            price=parse_number(parts[3] if len(parts) > 3 else None),
            open_price=parse_number(parts[1] if len(parts) > 1 else None),
            high=parse_number(parts[4] if len(parts) > 4 else None),
            low=parse_number(parts[5] if len(parts) > 5 else None),
            prev_close=parse_number(parts[2] if len(parts) > 2 else None),
            name=parts[0] if parts else None,
            source="sina",
            as_of=f"{parts[30]} {parts[31]}" if len(parts) > 31 else None,
            volume=parse_number(parts[8] if len(parts) > 8 else None),
            turnover=parse_number(parts[9] if len(parts) > 9 else None),
        )
    except (urllib.error.URLError, TimeoutError, http.client.IncompleteRead, IndexError) as exc:
        print(f"Sina fallback failed for {ticker}: {exc}", file=sys.stderr)
        return None


def fetch_fallback_quote(ticker: str):
    if ticker.startswith("US."):
        quote = fetch_yahoo_quote(ticker)
        if quote:
            return quote
    return fetch_sina_quote(ticker)


def fetch_quotes(tickers):
    futu_quotes, futu_failures = fetch_futu_quotes(tickers)
    quotes = dict(futu_quotes)
    quote_failures = {}

    for ticker in tickers:
        if ticker in quotes:
            continue
        fallback = fetch_fallback_quote(ticker)
        if fallback:
            quotes[ticker] = fallback
        else:
            quote_failures[ticker] = futu_failures.get(ticker, "quote unavailable")
        time.sleep(0.05)

    return quotes, quote_failures, futu_failures


def fetch_monthly_trends(tickers):
    trends = {}
    failures = {}
    for ticker in tickers:
      trend = fetch_yahoo_monthly_trend(ticker)
      if trend:
          trends[ticker] = trend
      else:
          failures[ticker] = "monthly trend unavailable"
      time.sleep(0.05)
    return trends, failures


def build_company(row_values, rank: int, quote=None, monthly_trend=None):
    raw_ticker = (row_values[0] or "").strip()
    if not raw_ticker:
        return None

    workbook_name = row_values[1] or normalize_ticker(raw_ticker)
    company_name = workbook_name or (quote or {}).get("name") or normalize_ticker(raw_ticker)
    if str(company_name).strip() == normalize_ticker(raw_ticker) and quote and quote.get("name"):
        company_name = quote["name"]

    price = (quote or {}).get("price", parse_number(row_values[2]))
    open_price = (quote or {}).get("open", parse_number(row_values[5]))
    prev_close = (quote or {}).get("prevClose", parse_number(row_values[6]))
    high = (quote or {}).get("high", parse_number(row_values[7]))
    low = (quote or {}).get("low", parse_number(row_values[8]))
    daily_change = None
    if quote and quote.get("dailyChange") is not None:
        daily_change = quote["dailyChange"]
    elif price is not None and prev_close not in (None, 0):
        daily_change = ((price - prev_close) / prev_close) * 100

    market = raw_ticker.split(".", 1)[0].upper()
    return {
        "rank": rank,
        "ticker": raw_ticker,
        "displayTicker": normalize_ticker(raw_ticker),
        "market": market,
        "name": str(company_name).strip() or normalize_ticker(raw_ticker),
        "price": price,
        "open": open_price,
        "prevClose": prev_close,
        "high": high,
        "low": low,
        "dailyChange": daily_change,
        "trend": build_snapshot_trend(price, open_price, high, low, prev_close),
        "monthlyTrend": monthly_trend or [],
        "quoteSource": (quote or {}).get("source", "workbook"),
        "quoteTime": (quote or {}).get("asOf"),
    }


def average_company_monthly_trends(companies):
    month_buckets = {}
    for company in companies:
        for point in company.get("monthlyTrend") or []:
            month = point.get("month")
            if not month:
                continue
            bucket = month_buckets.setdefault(month, {
                "month": month,
                "label": point.get("label") or month[5:],
                "open": [],
                "high": [],
                "low": [],
                "close": [],
            })
            for key in ("open", "high", "low", "close"):
                value = point.get(key)
                if is_finite_number(value):
                    bucket[key].append(value)

    averaged = []
    for month in sorted(month_buckets):
        bucket = month_buckets[month]

        def avg(key):
            values = bucket[key]
            return sum(values) / len(values) if values else None

        point = {
            "month": month,
            "label": bucket["label"],
            "open": avg("open"),
            "high": avg("high"),
            "low": avg("low"),
            "close": avg("close"),
        }
        if all(is_finite_number(point[key]) for key in ("open", "high", "low", "close")):
            averaged.append(point)

    return averaged[-6:]


def build_sector(sheet, quotes, monthly_trends):
    layer_code, sector_code, sector_name, sector_id = parse_sheet_meta(sheet.title)
    companies = []
    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
        raw_ticker = (row[0] or "").strip()
        company = build_company(row, idx, quotes.get(raw_ticker), monthly_trends.get(raw_ticker))
        if company:
            companies.append(company)

    valid_changes = [company["dailyChange"] for company in companies if company["dailyChange"] is not None]
    live_changes = [
        company["dailyChange"]
        for company in companies
        if company["dailyChange"] is not None and company.get("quoteSource") != "workbook"
    ]
    average_change = sum(valid_changes) / len(valid_changes) if valid_changes else 0.0
    monthly_trend = average_company_monthly_trends(companies)
    current_month_change = monthly_trend[-1]["close"] if monthly_trend else None
    movers = sorted(
        [company for company in companies if company["dailyChange"] is not None],
        key=lambda item: item["dailyChange"],
    )
    leader = movers[-1] if movers else None
    laggard = movers[0] if movers else None

    return layer_code, {
        "id": sector_id,
        "code": sector_code,
        "name": sector_name,
        "summary": f"{len(companies)} companies",
        "averageChange": average_change,
        "currentMonthChange": current_month_change if current_month_change is not None else average_change,
        "quoteCoverage": len(live_changes),
        "leader": {
            "ticker": leader["displayTicker"],
            "change": leader["dailyChange"],
        } if leader else None,
        "laggard": {
            "ticker": laggard["displayTicker"],
            "change": laggard["dailyChange"],
        } if laggard else None,
        "trend": average_company_trends(companies),
        "monthlyTrend": monthly_trend,
        "companies": companies,
    }


def extract_tickers(workbook):
    tickers = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(min_row=2, values_only=True):
            raw_ticker = (row[0] or "").strip()
            if raw_ticker and raw_ticker not in tickers:
                tickers.append(raw_ticker)
    return tickers


def build_industry_chain():
    workbook = load_workbook(WORKBOOK_PATH, data_only=True)
    tickers = extract_tickers(workbook)
    quotes, quote_failures, futu_failures = fetch_quotes(tickers)
    monthly_trends, monthly_failures = fetch_monthly_trends(tickers)

    grouped_layers = {
        layer_code: {
            "id": meta["id"],
            "code": layer_code,
            "name": meta["name"],
            "nameEn": meta["nameEn"],
            "sectors": [],
        }
        for layer_code, meta in LAYER_META.items()
    }

    for sheet in workbook.worksheets:
        layer_code, sector = build_sector(sheet, quotes, monthly_trends)
        grouped_layers[layer_code]["sectors"].append(sector)

    chain = [grouped_layers[layer_code] for layer_code in ("L0", "L1", "L2") if grouped_layers[layer_code]["sectors"]]
    source_counts = {}
    for quote in quotes.values():
        source = quote.get("source", "unknown")
        source_counts[source] = source_counts.get(source, 0) + 1

    metadata = {
        "sourceWorkbook": str(WORKBOOK_PATH),
        "quoteSyncedAt": datetime.now().astimezone().isoformat(timespec="seconds"),
        "quoteSource": "futu+fallback",
        "companyCount": len(tickers),
        "quoteCoverage": len(quotes),
        "monthlyCoverage": len(monthly_trends),
        "quoteSourceCounts": source_counts,
        "quoteMissing": quote_failures,
        "monthlyMissing": monthly_failures,
        "futuUnsupported": futu_failures,
    }
    return chain, metadata


def main():
    industry_chain, metadata = build_industry_chain()
    payload = json.dumps(industry_chain, ensure_ascii=False, indent=2)
    metadata_payload = json.dumps(metadata, ensure_ascii=False, indent=2)
    content = (
        "// Auto-generated by scripts/sync_l0_l2_board_data.py.\n"
        f"// Source workbook: {WORKBOOK_PATH}\n\n"
        f"export const DATA_SOURCE = {metadata_payload};\n\n"
        f"export const INDUSTRY_CHAIN = {payload};\n"
    )
    OUTPUT_PATH.write_text(content, encoding="utf-8")
    print(f"Wrote {OUTPUT_PATH}")
    print(
        "Quote coverage: "
        f"{metadata['quoteCoverage']}/{metadata['companyCount']} "
        f"via {metadata['quoteSourceCounts']}"
    )
    if metadata["quoteMissing"]:
        print(f"Missing quotes: {metadata['quoteMissing']}")


if __name__ == "__main__":
    main()
